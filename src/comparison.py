from __future__ import annotations

from pathlib import Path
from typing import Any
import math

import pandas as pd

from .storage import BASE_DIR, EXPORTS_DIR, SESSIONS_DIR, read_json, session_path
from .session_exports import latest_csv_by_role, infer_video_role_from_name, export_session_debug_files


def _num(value: Any) -> float | None:
    if value in (None, "", [], {}):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _round(value: Any, digits: int = 3) -> Any:
    n = _num(value)
    if n is None:
        return value if value not in (None, []) else ""
    return round(n, digits)


def _is_blank(value: Any) -> bool:
    return value in ("", None, [], {})


def _display_value(value: Any, fallback: str = "N/A") -> Any:
    if _is_blank(value):
        return fallback
    return _round(value)


def _display_side(value: Any, sk_value: Any = None) -> str:
    if not _is_blank(value):
        return str(value)
    if not _is_blank(sk_value):
        return "bilateral/전체"
    return "N/A"


def _explicit_cell(value: Any, fallback: str = "N/A") -> Any:
    """Customer-facing export must not contain visually blank cells."""
    if value is None:
        return fallback
    try:
        if isinstance(value, float) and math.isnan(value):
            return fallback
    except Exception:
        pass
    if value == "" or value == [] or value == {}:
        return fallback
    return value


def _explicit_row(row: dict[str, Any]) -> dict[str, Any]:
    return {k: _explicit_cell(v) for k, v in row.items()}


def _mean(values: list[Any]) -> float | None:
    nums = [_num(v) for v in values]
    nums = [v for v in nums if v is not None]
    if not nums:
        return None
    return sum(nums) / len(nums)


def _diff(a: Any, b: Any) -> Any:
    na, nb = _num(a), _num(b)
    if na is None or nb is None:
        return ""
    return round(na - nb, 3)


def _diff_percent(a: Any, b: Any) -> Any:
    na, nb = _num(a), _num(b)
    if na is None or nb is None or abs(nb) < 1e-9:
        return ""
    return round(abs(na - nb) / abs(nb) * 100.0, 1)


def _within_10_percent(a: Any, b: Any) -> str:
    pct = _diff_percent(a, b)
    if pct == "":
        return ""
    return "Y" if float(pct) <= 10.0 else "N"


def _absolute_error(a: Any, b: Any) -> Any:
    na, nb = _num(a), _num(b)
    if na is None or nb is None:
        return ""
    return round(abs(na - nb), 3)


def _metric_tolerance(metric: str, unit: str) -> float | None:
    # v0.5.9: clinical/running comparison should not use percent error alone.
    # Angles near zero can show huge percent errors despite being clinically close.
    tolerances = {
        "Forward Lean": 3.0,
        "Shank Angle @ touch-down": 3.0,
        "Pelvic Drop": 3.0,
        "Trunk Lateral Tilt": 3.0,
        "Rear Shoulder Tilt": 3.0,
        "Rear Knee Alignment @ mid-stance": 3.0,
        "Rear Knee Alignment Left @ mid-stance": 3.0,
        "Rear Knee Alignment Right @ mid-stance": 3.0,
        "Max Thigh Flexion": 5.0,
        "Max Thigh Extension": 5.0,
        "Hip ROM": 5.0,
        "Knee Flexion @ touch-down": 5.0,
        "Max Knee Flexion @ stance": 5.0,
        "Max Knee Flexion @ swing": 5.0,
        "Knee ROM": 5.0,
        "Cadence": 10.0,
        "Contact Time": 0.04,
        "Overstride": 15.0,
        "Vertical Displacement": 15.0,
        "Step Separation": 15.0,
    }
    return tolerances.get(metric)


def _within_metric_tolerance(metric: str, a: Any, b: Any, unit: str) -> str:
    pct = _diff_percent(a, b)
    abs_err = _absolute_error(a, b)
    tol = _metric_tolerance(metric, unit)
    if pct == "" and abs_err == "":
        return ""
    if pct != "" and float(pct) <= 10.0:
        return "Y"
    if tol is not None and abs_err != "" and float(abs_err) <= tol:
        return "Y"
    return "N"


def _issue_group(metric: str, pct: Any, unit: str, clip: dict[str, Any]) -> str:
    if pct == "":
        return ""
    try:
        if float(pct) <= 10.0:
            return "10% 이내"
    except Exception:
        return ""
    m = str(metric)
    if m in {"Cadence", "Contact Time", "Shank Angle @ touch-down", "Knee Flexion @ touch-down", "Overstride"}:
        return "10% 초과 - 이벤트/touch-down 보정 필요"
    if m in {"Max Knee Flexion @ stance", "Max Knee Flexion @ swing", "Knee ROM"}:
        return "10% 초과 - stance/swing phase 보정 필요"
    if m in {"Vertical Displacement", "Step Separation"}:
        return "10% 초과 - px→mm 스케일 보정 필요"
    if m in {"Max Thigh Extension"}:
        return "10% 초과 - 부호/대퇴 extension 기준 확인"
    if m in {"Max Thigh Flexion", "Hip ROM"}:
        return "10% 초과 - 대퇴각 산출축/ROM 기준 확인"
    if m in {"Pelvic Drop", "Trunk Lateral Tilt", "Rear Shoulder Tilt", "Knee Alignment @ mid-stance", "Rear Knee Alignment @ mid-stance", "Rear Knee Alignment Left @ mid-stance", "Rear Knee Alignment Right @ mid-stance"}:
        return "10% 초과 - 후면 축/부호/측정정의 확인"
    return "10% 초과 - 보정 필요"


def _value(values: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        val = values.get(key)
        if val not in (None, "", [], {}):
            return val
    return ""


def compute_auto_motionmetrix_values(values: dict[str, Any]) -> dict[str, Any]:
    """Compute customer-requested derived MotionMetrix fields.

    Customers enter only source values such as flexion/extension or landing/swing max.
    ROM, overall average, and left-right difference are computed automatically here.
    The function also keeps backward compatibility with older field names.
    """
    out = dict(values or {})

    # Backward-compatible unit normalization for older exports.
    over_cm = _num(out.get("overstride_mean_cm"))
    if "overstride_mean_mm" not in out and over_cm is not None:
        out["overstride_mean_mm"] = round(over_cm * 10.0, 3)

    vo_cm = _num(out.get("vertical_oscillation_mean_cm"))
    if "vertical_oscillation_mean_mm" not in out and vo_cm is not None:
        out["vertical_oscillation_mean_mm"] = round(vo_cm * 10.0, 3)

    raw_cadence = _num(out.get("cadence_raw_value"))
    raw_unit = out.get("cadence_raw_unit")
    if "cadence_steps_per_min" not in out and raw_cadence is not None:
        if raw_unit == "strides/min":
            out["cadence_steps_per_min"] = round(raw_cadence * 2.0, 3)
        elif raw_unit == "steps/min":
            out["cadence_steps_per_min"] = round(raw_cadence, 3)

    contact_sec = _num(out.get("contact_time_mean_sec"))
    contact_ms = _num(out.get("contact_time_mean_ms"))
    # v0.5.7: MotionMetrix screen shows Contact Time in seconds (e.g. 0.225 s).
    # The UI field is legacy ms. If a user enters 0.225 into that field, treat it
    # as seconds instead of exporting 0.000225 s rounded to 0.0.
    if contact_sec is None and contact_ms is not None and 0 < contact_ms <= 5:
        contact_sec = contact_ms
        contact_ms = contact_sec * 1000.0
        out["contact_time_input_interpreted_as_sec"] = True
    if contact_ms is None and contact_sec is not None:
        contact_ms = contact_sec * 1000.0
    if contact_sec is None and contact_ms is not None:
        contact_sec = contact_ms / 1000.0
    if contact_ms is not None:
        out["contact_time_mean_ms"] = round(contact_ms, 3)
    if contact_sec is not None:
        out["contact_time_mean_sec"] = round(contact_sec, 3)

    # Hip/thigh ROM: customer confirmed flexion and extension are summed as magnitudes.
    # Example: flexion 20 deg + extension 20 deg = ROM 40 deg.
    # This works whether extension is entered as +20 or -20.
    lf = _num(out.get("left_thigh_flexion_max_deg"))
    le = _num(out.get("left_thigh_extension_max_deg"))
    rf = _num(out.get("right_thigh_flexion_max_deg"))
    re = _num(out.get("right_thigh_extension_max_deg"))
    left_hip_rom = (abs(lf) + abs(le)) if lf is not None and le is not None else _num(out.get("left_hip_rom_deg"))
    right_hip_rom = (abs(rf) + abs(re)) if rf is not None and re is not None else _num(out.get("right_hip_rom_deg"))
    if left_hip_rom is not None:
        out["left_hip_rom_deg"] = round(left_hip_rom, 3)
    if right_hip_rom is not None:
        out["right_hip_rom_deg"] = round(right_hip_rom, 3)
    hip_mean = _mean([left_hip_rom, right_hip_rom])
    if hip_mean is not None:
        out["hip_rom_mean_deg"] = round(hip_mean, 3)
    if left_hip_rom is not None and right_hip_rom is not None:
        out["hip_rom_asymmetry_deg"] = round(abs(left_hip_rom - right_hip_rom), 3)

    # MotionMetrix Gait Characteristics table means from left/right inputs.
    thigh_flex_mean = _mean([out.get("left_thigh_flexion_max_deg"), out.get("right_thigh_flexion_max_deg")])
    if thigh_flex_mean is not None:
        out["thigh_flexion_mean_deg"] = round(thigh_flex_mean, 3)
    thigh_ext_mean = _mean([out.get("left_thigh_extension_max_deg"), out.get("right_thigh_extension_max_deg")])
    if thigh_ext_mean is not None:
        out["thigh_extension_mean_deg"] = round(thigh_ext_mean, 3)

    # Knee ROM: swing max flexion minus landing flexion.
    ll = _num(out.get("left_knee_flexion_landing_deg"))
    ls = _num(out.get("left_knee_flexion_swing_max_deg"))
    rl = _num(out.get("right_knee_flexion_landing_deg"))
    rs = _num(out.get("right_knee_flexion_swing_max_deg"))
    left_knee_rom = abs(ls - ll) if ls is not None and ll is not None else _num(out.get("left_knee_rom_deg"))
    right_knee_rom = abs(rs - rl) if rs is not None and rl is not None else _num(out.get("right_knee_rom_deg"))
    if left_knee_rom is not None:
        out["left_knee_rom_deg"] = round(left_knee_rom, 3)
    if right_knee_rom is not None:
        out["right_knee_rom_deg"] = round(right_knee_rom, 3)
    knee_mean = _mean([left_knee_rom, right_knee_rom])
    if knee_mean is not None:
        out["knee_rom_mean_deg"] = round(knee_mean, 3)
    if left_knee_rom is not None and right_knee_rom is not None:
        out["knee_rom_asymmetry_deg"] = round(abs(left_knee_rom - right_knee_rom), 3)
    landing_mean = _mean([ll, rl])
    if landing_mean is not None:
        out["knee_flexion_landing_mean_deg"] = round(landing_mean, 3)
    stance_mean = _mean([out.get("left_knee_flexion_stance_max_deg"), out.get("right_knee_flexion_stance_max_deg")])
    if stance_mean is not None:
        out["knee_flexion_stance_max_mean_deg"] = round(stance_mean, 3)
    swing_mean = _mean([out.get("left_knee_flexion_swing_max_deg"), out.get("right_knee_flexion_swing_max_deg")])
    if swing_mean is not None:
        out["knee_flexion_swing_max_mean_deg"] = round(swing_mean, 3)

    # Backward-compatible averages for rear/side metrics.
    shank_mean = _mean([out.get("left_shank_angle_signed_deg"), out.get("right_shank_angle_signed_deg")])
    if "shank_angle_mean_signed_deg" not in out and shank_mean is not None:
        out["shank_angle_mean_signed_deg"] = round(shank_mean, 3)

    pelvic_mean = _mean([out.get("left_stance_pelvic_drop_deg"), out.get("right_stance_pelvic_drop_deg")])
    if "pelvic_drop_mean_deg" not in out and pelvic_mean is not None:
        out["pelvic_drop_mean_deg"] = round(pelvic_mean, 3)

    trunk_mean = _mean([out.get("left_stance_trunk_lateral_tilt_deg"), out.get("right_stance_trunk_lateral_tilt_deg")])
    if "trunk_lateral_tilt_mean_deg" not in out and trunk_mean is not None:
        out["trunk_lateral_tilt_mean_deg"] = round(trunk_mean, 3)

    kcollapse_mean = _mean([out.get("left_knee_medial_collapse_mean"), out.get("right_knee_medial_collapse_mean")])
    if "knee_medial_collapse_mean" not in out and kcollapse_mean is not None:
        out["knee_medial_collapse_mean"] = round(kcollapse_mean, 3)

    step_cm = _num(out.get("step_width_mean_cm"))
    if "step_width_mean_mm" not in out and step_cm is not None:
        out["step_width_mean_mm"] = round(step_cm * 10.0, 3)

    return out


def _data_mtime(path: Path) -> float:
    try:
        return path.stat().st_mtime
    except Exception:
        return 0.0


def _read_csv_first_row(path: Path) -> dict[str, Any] | None:
    try:
        df = pd.read_csv(path)
        if df.empty:
            return None
        data = df.iloc[0].to_dict()
        return data
    except Exception:
        return None


def _read_csv_all(path: Path) -> pd.DataFrame | None:
    try:
        df = pd.read_csv(path)
        if df.empty:
            return None
        return df
    except Exception:
        return None


def _safe_mean(series: Any, abs_value: bool = False) -> Any:
    try:
        nums = pd.to_numeric(series, errors="coerce").dropna()
        if nums.empty:
            return ""
        if abs_value:
            nums = nums.abs()
        return round(float(nums.mean()), 3)
    except Exception:
        return ""


def _safe_range(series: Any) -> Any:
    try:
        nums = pd.to_numeric(series, errors="coerce").dropna()
        if nums.empty:
            return ""
        return round(float(nums.max() - nums.min()), 3)
    except Exception:
        return ""


def _fallback_clip_summary_from_frame_csv(path: Path) -> dict[str, Any] | None:
    """Build a minimal clip summary when a processed clip_summary CSV is missing.

    This is intentionally conservative. It mainly prevents the final comparison
    sheet from showing blanks when frame_metrics/frame_stats already contain
    usable rear or side skeleton values. Event-specific values are only filled
    if corresponding columns already exist in the frame CSV.
    """
    df = _read_csv_all(path)
    if df is None or df.empty:
        return None
    view = str(df.get("view_type", pd.Series(["unknown"])).dropna().iloc[0] if "view_type" in df else "unknown")
    data: dict[str, Any] = {
        "view_type": view,
        "valid_frame_count": int(len(df)),
        "_clip_summary_path": "",
        "_fallback_frame_csv_path": str(path.relative_to(BASE_DIR)) if path.is_relative_to(BASE_DIR) else str(path),
        "_source_video": str(df.get("source_video_name", pd.Series([""])).dropna().iloc[0]) if "source_video_name" in df and not df.get("source_video_name", pd.Series()).dropna().empty else "",
        "_created_at": path.stem,
        "_source_mtime": _data_mtime(path),
        "summary_source": "fallback_from_frame_csv",
    }
    if "source_fps" in df:
        data["actual_video_fps"] = _safe_mean(df["source_fps"])
    if "timestamp_sec" in df:
        times = pd.to_numeric(df["timestamp_sec"], errors="coerce").dropna()
        if not times.empty:
            data["valid_duration_sec"] = round(float(times.max() - times.min()), 3)
    if "pose_detected" in df:
        detected = df["pose_detected"].astype(str).str.lower().isin(["true", "1", "yes"])
        data["pose_detection_rate"] = round(float(detected.mean()), 3)

    # Side values that can be safely summarized from frame rows.
    if "forward_lean_deg" in df:
        v = _safe_mean(df["forward_lean_deg"])
        data["forward_lean_signed_avg_deg"] = v
        try:
            data["forward_lean_avg_deg"] = round(abs(float(v)), 3)
        except Exception:
            data["forward_lean_avg_deg"] = v
    if "pelvis_center_y" in df:
        data["pelvis_vertical_oscillation_px"] = _safe_range(df["pelvis_center_y"])
    if "pelvis_vertical_displacement_mm_est" in df:
        data["pelvis_vertical_displacement_mm_est"] = _safe_mean(df["pelvis_vertical_displacement_mm_est"])
    elif "pelvis_vertical_oscillation_mm_est" in df:
        data["pelvis_vertical_displacement_mm_est"] = _safe_mean(df["pelvis_vertical_oscillation_mm_est"])

    # Rear values used by final comparison.
    if "rear_pelvic_tilt_deg" in df:
        data["rear_pelvic_tilt_avg_deg"] = _safe_mean(df["rear_pelvic_tilt_deg"])
    if "rear_trunk_lateral_tilt_deg" in df:
        data["rear_trunk_lateral_tilt_avg_deg"] = _safe_mean(df["rear_trunk_lateral_tilt_deg"])
    if "step_width_mm_est" in df:
        data["step_width_avg_mm_est"] = _safe_mean(df["step_width_mm_est"])
    if "step_width_px" in df:
        data["step_width_avg_px"] = _safe_mean(df["step_width_px"])
    km_cols = [c for c in ["left_knee_medial_offset_px", "right_knee_medial_offset_px"] if c in df]
    if km_cols:
        vals = pd.concat([pd.to_numeric(df[c], errors="coerce") for c in km_cols], ignore_index=True).dropna()
        data["knee_medial_collapse_avg_px"] = round(float(vals.abs().mean()), 3) if not vals.empty else ""

    return data


def _latest_clip_summaries(session_id: str) -> dict[str, dict[str, Any]]:
    """Return latest clip summaries by exact video role.

    v0.5.7 intentionally stops falling back from side_running metrics to
    side_static files. Each MotionMetrix comparison row declares its required
    source role, so a static forward-lean CSV cannot become the source for
    cadence/contact/ROM values.
    """
    result: dict[str, dict[str, Any]] = {}

    for role, item in latest_csv_by_role(session_id, "clip_summary").items():
        data = _read_csv_first_row(item["path"])
        if not data:
            continue
        data["_clip_summary_path"] = item["relative_path"]
        data["_created_at"] = item.get("created_at", "") or item["path"].stem
        data["_source_video"] = item.get("source_video", "") or data.get("source_video_name", "")
        data["_source_mtime"] = item.get("mtime", 0)
        data["summary_source"] = "clip_summary_csv"
        data["video_role"] = role
        result[role] = data

    # Conservative fallback: only from frame CSV of the same exact role.
    # This may fill rear skeleton-only static quantities, but will not cross map
    # side_static into side_running.
    for kind in ["frame_metrics", "frame_stats"]:
        for role, item in latest_csv_by_role(session_id, kind).items():
            if role in result and result[role].get("summary_source") == "clip_summary_csv":
                continue
            data = _fallback_clip_summary_from_frame_csv(item["path"])
            if not data:
                continue
            data["video_role"] = role
            data["summary_source"] = f"fallback_from_{kind}"
            result[role] = data
    return result



def _clip_event_score(clip: dict[str, Any], metric: str, values: dict[str, Any]) -> float:
    if not clip:
        return -1e9
    score = 0.0
    event_count = _num(clip.get("event_count_used")) or 0.0
    duration = _num(clip.get("valid_duration_sec")) or 0.0
    left = _num(clip.get("left_step_count")) or 0.0
    right = _num(clip.get("right_step_count")) or 0.0
    pose_rate = _num(clip.get("pose_detection_rate")) or 0.0
    score += pose_rate * 10.0
    if event_count >= 6:
        score += 5.0
    if left + right > 0:
        score += (1.0 - abs(left - right) / max(left + right, 1.0)) * 5.0
    cadence = _num(clip.get("estimated_cadence_spm"))
    mm_cadence = _num(values.get("cadence_steps_per_min"))
    if metric == "Cadence" and cadence is not None:
        if mm_cadence is not None:
            score += max(0.0, 20.0 - abs(cadence - mm_cadence) / 5.0)
        elif 140 <= cadence <= 210:
            score += 10.0
        else:
            score -= 5.0
    if metric == "Contact Time":
        ct = _num(clip.get("contact_time_avg_sec"))
        mm_ct = _num(values.get("contact_time_mean_sec"))
        if ct is not None and mm_ct is not None:
            score += max(0.0, 12.0 - abs(ct - mm_ct) / 0.02)
        if event_count >= 4:
            score += 3.0
    if str(clip.get("timing_confidence", "")).startswith("low"):
        score -= 1.0
    return score


def _select_clip_for_spec(spec: dict[str, Any], clips: dict[str, dict[str, Any]], values: dict[str, Any]) -> tuple[dict[str, Any], str, str]:
    metric = spec.get("metric", "")
    source_role = spec.get("source_role", spec.get("view", ""))
    # v0.5.9: cadence/contact time are event-level quantities. Side view can
    # miss the far foot; rear view often captures left/right alternation better.
    if metric in {"Cadence", "Contact Time"}:
        candidates = []
        for role in ["rear_running", "side_running"]:
            if role in clips:
                candidates.append((role, clips[role], _clip_event_score(clips[role], metric, values)))
        if candidates:
            role, clip, _score = max(candidates, key=lambda x: x[2])
            return clip, role, "best_source_by_event_quality"
    return clips.get(source_role, {}) if source_role in ("side_running", "side_static", "rear_running", "rear_static") else {}, source_role, "fixed_source_role"


COMPARISON_SPECS = [
    # Running Performance page
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Running Economy", "sk":"", "sk_unit":"", "mm":"running_economy_j_kg_m", "mm_unit":"J/kg/m", "target_only":True, "source":"MotionMetrix target", "note":"3차에서 여러 Skeleton feature로 예측할 MotionMetrix 정답값입니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Cadence", "sk":"estimated_cadence_spm", "sk_unit":"/min", "mm":"cadence_steps_per_min", "mm_unit":"/min", "source":"initial contact event count / valid duration", "note":"MotionMetrix 화면의 /min 표기와 맞춰 표시합니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Contact Time", "sk":"contact_time_avg_sec", "sk_unit":"s", "mm":"contact_time_mean_sec", "mm_unit":"s", "caution":True, "source":"contact start-to-toe-off event average", "note":"실제 영상 FPS와 지면선/접촉 판별에 민감합니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Forward Lean", "sk":"forward_lean_avg_deg", "sk_unit":"deg", "mm":"forward_lean_deg", "mm_unit":"deg", "source":"valid side frames / abs signed lean for MotionMetrix-style display", "note":"진행방향 부호로 계산 후 MotionMetrix 화면 비교용으로 절대값 평균을 표시합니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Overstride", "sk":"overstride_selected_mm_est", "sk_unit":"mm", "mm":"overstride_mean_mm", "mm_unit":"mm", "caution":True, "source":"initial contact pelvis projection-to-landing ankle distance", "note":"MediaPipe 추정 좌표/스케일 기반입니다. MotionMetrix depth camera 계측값과 완전 동일한 값은 아닙니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Vertical Displacement", "sk":"pelvis_vertical_displacement_mm_est", "sk_unit":"mm", "mm":"vertical_oscillation_mean_mm", "mm_unit":"mm", "caution":True, "source":"image pelvis vertical range + height-based px-to-mm estimate", "note":"MotionMetrix의 COM vertical range 정의에 맞춰 평균이 아닌 range로 계산합니다."},
    {"screen":"Running Performance", "view":"side", "source_role":"side_running", "metric":"Braking Force (max)", "sk":"", "sk_unit":"", "mm":"braking_force_mean_value", "mm_unit":"braking_force_unit", "target_only":True, "source":"MotionMetrix target", "note":"Skeleton으로 직접 계산하지 않고 3차 학습 정답값으로 사용합니다."},
    {"screen":"Running Performance", "view":"aggregate", "metric":"Vertical Force (max)", "sk":"", "sk_unit":"", "mm":"vertical_force_max_optional", "mm_unit":"BW", "target_only":True, "source":"MotionMetrix optional input"},
    {"screen":"Running Performance", "view":"aggregate", "metric":"Lateral Force (max)", "sk":"", "sk_unit":"", "mm":"lateral_force_max_optional", "mm_unit":"Fv", "target_only":True, "source":"MotionMetrix optional input"},
    {"screen":"Running Performance", "view":"aggregate", "metric":"Stride Rating", "sk":"", "sk_unit":"", "mm":"stride_rating_optional", "mm_unit":"score", "target_only":True, "source":"MotionMetrix optional input"},

    # Gait Characteristics page
    {"screen":"Gait Characteristics", "view":"rear", "source_role":"rear_running", "metric":"Step Separation", "sk":"step_width_avg_mm_est", "sk_unit":"mm", "mm":"step_width_mean_mm", "mm_unit":"mm", "caution":True, "skeleton_only_when_missing":True, "source":"rear/ankle separation estimate", "note":"MotionMetrix 화면의 Step Separation에 대응합니다. 후면 MotionMetrix 직접 입력이 없으면 Skeleton-only로 표시됩니다."},
    {"screen":"Gait Characteristics", "view":"rear", "source_role":"rear_running", "metric":"Knee Alignment @ mid-stance", "sk":"knee_alignment_rear_abs_mean_deg", "sk_unit":"deg", "mm":"knee_medial_collapse_mean", "mm_unit":"deg", "caution":True, "skeleton_only_when_missing":True, "source":"rear 2D hip-knee-ankle frontal-plane angle proxy", "note":"후면에서 산출 가능한 무릎 정렬 각도입니다. 단일 RGB 기반 Skeleton 각도이므로 MotionMetrix depth 계측값과 완전 동일하다고 보지는 않습니다."},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Max Thigh Flexion", "sk":"max_thigh_flexion_mean_deg", "sk_unit":"deg", "mm":"thigh_flexion_mean_deg", "mm_unit":"deg", "source":"thigh vector vs vertical / cycle max average"},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Max Thigh Extension", "sk":"max_thigh_extension_mean_deg", "sk_unit":"deg", "mm":"thigh_extension_mean_deg", "mm_unit":"deg", "source":"thigh vector vs vertical / MotionMetrix-style signed extension"},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Hip ROM", "sk":"hip_rom_avg_deg", "sk_unit":"deg", "mm":"hip_rom_mean_deg", "mm_unit":"deg", "auto":True, "source":"Max Thigh Flexion + Max Thigh Extension", "note":"굴곡 20도 + 신전 20도 = ROM 40도 기준입니다."},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Shank Angle @ touch-down", "sk":"shank_angle_at_contact_selected_avg_deg", "sk_unit":"deg", "mm":"shank_angle_mean_signed_deg", "mm_unit":"deg", "source":"initial contact event / shank vector vs vertical"},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Knee Flexion @ touch-down", "sk":"knee_flexion_touchdown_avg_deg", "sk_unit":"deg", "mm":"knee_flexion_landing_mean_deg", "mm_unit":"deg", "source":"initial contact event / 180 - included knee angle", "note":"관절 내각이 아니라 MotionMetrix와 같은 굴곡각 기준입니다."},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Max Knee Flexion @ stance", "sk":"knee_flexion_stance_max_mean_deg", "sk_unit":"deg", "mm":"knee_flexion_stance_max_mean_deg", "mm_unit":"deg", "source":"stance phase max knee flexion"},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Max Knee Flexion @ swing", "sk":"knee_flexion_swing_max_mean_deg", "sk_unit":"deg", "mm":"knee_flexion_swing_max_mean_deg", "mm_unit":"deg", "source":"swing phase max knee flexion"},
    {"screen":"Gait Characteristics", "view":"side", "source_role":"side_running", "metric":"Knee ROM", "sk":"knee_rom_avg_deg", "sk_unit":"deg", "mm":"knee_rom_mean_deg", "mm_unit":"deg", "auto":True, "source":"Max Knee Flexion @ swing - Knee Flexion @ touch-down"},

    # Skeleton-only and manual/reference items retained from v0.5.4 UI
    # v0.5.12: explicitly share rear-view angles that Skeleton can calculate.
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Pelvic Drop", "sk":"rear_pelvic_tilt_avg_deg", "sk_unit":"deg", "mm":"pelvic_drop_mean_deg", "mm_unit":"deg", "skeleton_only_when_missing":True, "source":"rear skeleton pelvis line tilt", "note":"후면에서 산출 가능한 각도입니다."},
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Trunk Lateral Tilt", "sk":"rear_trunk_lateral_tilt_avg_deg", "sk_unit":"deg", "mm":"trunk_lateral_tilt_mean_deg", "mm_unit":"deg", "skeleton_only_when_missing":True, "source":"rear skeleton trunk line tilt", "note":"후면에서 산출 가능한 각도입니다."},
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Rear Shoulder Tilt", "sk":"rear_shoulder_tilt_avg_deg", "sk_unit":"deg", "mm":"", "mm_unit":"", "skeleton_only":True, "source":"rear shoulder line tilt", "note":"후면 어깨선 좌우 기울기 Skeleton-only 참고값입니다."},
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Rear Knee Alignment Left @ mid-stance", "sk":"left_knee_alignment_rear_avg_deg", "sk_unit":"deg", "mm":"", "mm_unit":"", "skeleton_only":True, "source":"rear left hip-knee-ankle alignment angle proxy", "note":"후면 좌측 무릎 정렬각 Skeleton-only 참고값입니다."},
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Rear Knee Alignment Right @ mid-stance", "sk":"right_knee_alignment_rear_avg_deg", "sk_unit":"deg", "mm":"", "mm_unit":"", "skeleton_only":True, "source":"rear right hip-knee-ankle alignment angle proxy", "note":"후면 우측 무릎 정렬각 Skeleton-only 참고값입니다."},
    {"screen":"후면 Skeleton 각도", "view":"rear", "source_role":"rear_running", "metric":"Rear Knee Alignment @ mid-stance", "sk":"knee_alignment_rear_abs_mean_deg", "sk_unit":"deg", "mm":"", "mm_unit":"", "skeleton_only":True, "source":"rear bilateral knee alignment absolute mean angle proxy", "note":"후면 무릎 정렬각 좌우 절대 평균 Skeleton-only 참고값입니다."},
    {"screen":"Manual Label", "view":"side", "source_role":"side_running", "metric":"Strike Type", "sk":"foot_strike_type_summary", "sk_unit":"candidate", "mm":"session_representative_strike_type", "mm_unit":"manual", "manual_label":True, "source":"manual label + skeleton candidate", "note":"고객 수동 라벨입니다."},
    {"screen":"Aggregate", "view":"aggregate", "metric":"Running Type", "sk":"", "sk_unit":"", "mm":"running_type", "mm_unit":"class", "target_only":True, "source":"MotionMetrix target", "note":"3차 분류 모델 정답값입니다."},
]


def build_final_comparison_rows(session_id: str) -> list[dict[str, Any]]:
    values = read_json(session_path(session_id) / "motionmetrix_values.json", {})
    values = compute_auto_motionmetrix_values(values if isinstance(values, dict) else {})
    clips = _latest_clip_summaries(session_id)
    rows: list[dict[str, Any]] = []
    for spec in COMPARISON_SPECS:
        view = spec["view"]
        source_role = spec.get("source_role", view)
        clip, actual_source_role, source_selection_method = _select_clip_for_spec(spec, clips, values)
        sk_key = spec.get("sk", "")
        mm_key = spec.get("mm", "")
        sk_value = _value(clip, sk_key) if sk_key else ""
        mm_value = _value(values, mm_key) if mm_key else ""
        mm_unit_spec = spec.get("mm_unit", "")
        mm_unit = values.get(mm_unit_spec, "") if mm_unit_spec in values else mm_unit_spec
        sk_unit = spec.get("sk_unit", "")
        comparable = bool(sk_value not in ("", None) and mm_value not in ("", None) and sk_unit == mm_unit and sk_unit not in ("", "px", "raw", "ratio", "candidate", "manual", "class", "score", "BW", "Fv"))
        target_only = bool(spec.get("target_only"))
        skeleton_only = bool(spec.get("skeleton_only")) or (bool(spec.get("skeleton_only_when_missing")) and mm_value in ("", None))
        manual_label = bool(spec.get("manual_label"))
        caution = bool(spec.get("caution"))
        calc_status = clip.get("event_quality_summary", "") or ("source_missing" if source_role in ("side_running", "rear_running", "side_static", "rear_static") and not clip else "")
        # For thigh extension, customers may enter MotionMetrix extension as either
        # signed negative or positive magnitude. Compare magnitudes for the 10% QA flag.
        if spec["metric"] == "Max Thigh Extension" and comparable:
            sk_for_compare = abs(_num(sk_value) or 0.0)
            mm_for_compare = abs(_num(mm_value) or 0.0)
        else:
            sk_for_compare = sk_value
            mm_for_compare = mm_value
        diff_pct = _diff_percent(sk_for_compare, mm_for_compare) if comparable else ""
        abs_err = _absolute_error(sk_for_compare, mm_for_compare) if comparable else ""
        within_10 = _within_metric_tolerance(spec["metric"], sk_for_compare, mm_for_compare, sk_unit) if comparable else ""
        issue_group = _issue_group(spec["metric"], diff_pct, sk_unit, clip) if comparable else ""
        if comparable and within_10 == "Y" and diff_pct != "" and float(diff_pct) > 10.0:
            issue_group = "절대오차 허용범위 이내"
        timing_conf = clip.get("timing_confidence", "")
        low_fps = bool(str(timing_conf).startswith("low") or clip.get("low_fps_warning"))
        if spec["metric"] == "Overstride":
            scale_conf = clip.get("x_scale_confidence", clip.get("scale_confidence", ""))
        else:
            scale_conf = clip.get("scale_confidence", "")
        event_metrics = {"Cadence", "Contact Time", "Shank Angle @ touch-down", "Knee Flexion @ touch-down", "Overstride"}
        scale_metrics = {"Vertical Displacement", "Step Separation", "Overstride"}
        if target_only:
            status = "MotionMetrix 입력값 / 3차 학습 정답" if not _is_blank(mm_value) else "MotionMetrix target-only / 입력값 없음"
        elif source_role in ("side_running", "rear_running", "side_static", "rear_static") and not clip:
            status = f"Skeleton source missing: {source_role}"
        elif skeleton_only:
            status = "Skeleton-only / 후면 산출 가능 각도" if spec.get("screen") == "후면 Skeleton 각도" else "Skeleton-only / MotionMetrix 입력값 없음"
        elif manual_label:
            status = "수동 라벨 / 비교 대상 아님"
        elif comparable:
            if within_10 == "Y":
                status = "비교 가능 / 허용범위 이내"
            else:
                status = issue_group or "10% 초과 - 보정 필요"
            if caution and "추정" not in status:
                status += " / 추정값 주의"
            if low_fps and spec["metric"] in event_metrics:
                status += " / 저FPS 주의"
            if scale_conf in {"low", "unavailable"} and spec["metric"] in scale_metrics:
                status += " / 스케일 신뢰도 낮음"
        elif sk_value in ("", None) and mm_value in ("", None):
            status = "값 없음"
        elif sk_value in ("", None):
            status = "Skeleton 평균값 없음"
        elif mm_value in ("", None):
            status = "MotionMetrix 입력 필요"
        else:
            status = "단위 상이 또는 proxy 기준"
        if calc_status in ("event_count_low", "event_not_detected") and sk_key in {"estimated_cadence_spm", "contact_time_avg_sec", "overstride_avg_mm_est", "overstride_selected_mm_est", "shank_angle_at_contact_avg_deg", "shank_angle_at_contact_selected_avg_deg", "knee_flexion_touchdown_avg_deg"}:
            status = f"이벤트 검출 확인 필요 / {calc_status}" + (" / 저FPS 주의" if low_fps else "")

        raw_value = ""
        adjusted_value = ""
        selection_reason = ""
        source_side = ""
        if spec["metric"] == "Shank Angle @ touch-down":
            raw_value = clip.get("shank_angle_at_contact_raw_avg_deg", "")
            adjusted_value = clip.get("shank_angle_at_contact_corrected_avg_deg", "")
            selection_reason = clip.get("shank_angle_selection_reason", "")
            source_side = clip.get("shank_angle_selected_side", "")
        elif spec["metric"] == "Overstride":
            raw_value = clip.get("overstride_selected_px", "") or clip.get("overstride_avg_px", "")
            adjusted_value = clip.get("overstride_trimmed_mean_mm_est", "")
            x_src = clip.get("x_scale_source", "") or "N/A"
            x_conf = clip.get("x_scale_confidence", "") or "N/A"
            selection_reason = (clip.get("overstride_selection_reason", "") + f"; x_scale={x_src}/{x_conf}").strip('; ')
            source_side = clip.get("overstride_selected_side", "")
        elif spec["metric"] == "Cadence":
            raw_value = clip.get("cadence_count_spm", "")
            adjusted_value = clip.get("cadence_edge_adjusted_spm", "")
            selection_reason = clip.get("cadence_selection_method", "")
        elif spec["metric"] == "Contact Time":
            raw_value = clip.get("contact_time_avg_ms", "")
            adjusted_value = clip.get("contact_time_avg_sec", "")
            selection_reason = "low_fps_plateau_endpoint_adjusted" if low_fps else "contact_event_average"
        elif spec["metric"] in {"Vertical Displacement", "Step Separation"}:
            raw_value = clip.get("pelvis_vertical_oscillation_px", "") if spec["metric"] == "Vertical Displacement" else clip.get("step_width_avg_px", "")
            adjusted_value = sk_value
            selection_reason = f"scale_source={clip.get('scale_source','')}; confidence={clip.get('scale_confidence','')}"
        elif spec["metric"] in {"Max Thigh Flexion", "Max Thigh Extension", "Hip ROM", "Max Knee Flexion @ swing", "Knee ROM"}:
            raw_value = ""
            adjusted_value = sk_value
            selection_reason = "robust_percentile_or_cycle_median"

        # v0.5.11: make blanks explicit in the customer-facing comparison table.
        if target_only and _is_blank(sk_value):
            sk_display = "Skeleton 직접 산출 대상 아님"
        elif _is_blank(sk_value):
            sk_display = "Skeleton 없음"
        else:
            sk_display = _round(sk_value)
        raw_display = _display_value(raw_value, "N/A")
        adjusted_display = _display_value(adjusted_value, "보정 없음")
        mm_display = _round(mm_value) if not _is_blank(mm_value) else ("MotionMetrix 미입력" if mm_key else "N/A")
        source_side_display = _display_side(source_side, sk_value)

        row_out = {
            "session_id": session_id,
            "MotionMetrix 화면": spec.get("screen", ""),
            "구분": {"side":"측면", "rear":"후면", "aggregate":"종합"}.get(view, view),
            "측정 항목": spec["metric"],
            "Skeleton 평균값": sk_display,
            "Skeleton raw/audit 값": raw_display,
            "Skeleton adjusted 값": adjusted_display,
            "Skeleton 선택 사유": selection_reason or "N/A",
            "source_side": source_side_display,
            "Skeleton 단위": sk_unit or "N/A",
            "MotionMetrix 값": mm_display,
            "MotionMetrix 단위": mm_unit or "N/A",
            "차이값(Skeleton-MM)": _diff(sk_for_compare, mm_for_compare) if comparable else "",
            "절대오차": abs_err,
            "허용오차 기준": _metric_tolerance(spec["metric"], sk_unit) if comparable else "",
            "차이율(%)": diff_pct,
            "허용범위 여부": within_10,
            "문제 유형": issue_group,
            "비교 상태": status,
            "Skeleton 계산 방식": spec.get("source", ""),
            "required_video_role": source_role,
            "actual_video_role": clip.get("video_role", ""),
            "selected_source_role": actual_source_role,
            "source_selection_method": source_selection_method,
            "valid_frame_count": clip.get("valid_frame_count", ""),
            "valid_event_count": clip.get("event_count_used", ""),
            "raw_event_count": clip.get("event_count_raw", ""),
            "expected_event_count_from_MM": clip.get("expected_event_count_from_mm", ""),
            "actual_video_fps": clip.get("actual_video_fps", ""),
            "analysis_fps": clip.get("analysis_fps", ""),
            "timing_confidence": clip.get("timing_confidence", ""),
            "low_fps_warning": clip.get("low_fps_warning", ""),
            "scale_source": clip.get("scale_source", ""),
            "scale_confidence": clip.get("scale_confidence", ""),
            "x_scale_source": clip.get("x_scale_source", ""),
            "x_scale_confidence": clip.get("x_scale_confidence", ""),
            "y_scale_source": clip.get("y_scale_source", ""),
            "y_scale_confidence": clip.get("y_scale_confidence", ""),
            "calculation_status": calc_status,
            "Skeleton source CSV": clip.get("_clip_summary_path") or clip.get("_fallback_frame_csv_path", ""),
            "source_video": clip.get("_source_video", ""),
            "summary_source": clip.get("summary_source", ""),
            "비고": spec.get("note", "") + (" / 자동계산" if spec.get("auto") else ""),
        }
        rows.append(_explicit_row(row_out))
    return rows


def export_final_comparison_summary(session_ids: list[str] | None = None) -> list[Path]:
    rows: list[dict[str, Any]] = []
    if not SESSIONS_DIR.exists():
        return []
    if session_ids:
        target_sessions = [session_path(sid) for sid in session_ids]
    else:
        target_sessions = [p for p in SESSIONS_DIR.iterdir() if p.is_dir()]
    for session_dir in target_sessions:
        if session_dir.is_dir():
            rows.extend(build_final_comparison_rows(session_dir.name))
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.fillna("N/A").replace("", "N/A")
    csv_path = EXPORTS_DIR / "final_comparison_summary.csv"
    xlsx_path = EXPORTS_DIR / "final_comparison_summary.xlsx"
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    paths = [csv_path]
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Skeleton_vs_MotionMetrix")
            ws = writer.book["Skeleton_vs_MotionMetrix"]
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1F2937")
            header_font = Font(color="FFFFFF", bold=True)
            skeleton_fill = PatternFill("solid", fgColor="DBEAFE")
            mm_fill = PatternFill("solid", fgColor="FFEDD5")
            auto_fill = PatternFill("solid", fgColor="DCFCE7")
            warn_fill = PatternFill("solid", fgColor="FEF3C7")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            col_names = [cell.value for cell in ws[1]]
            for idx, name in enumerate(col_names, start=1):
                if name and "Skeleton" in str(name):
                    for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                        for c in cell:
                            c.fill = skeleton_fill
                if name and "MotionMetrix" in str(name):
                    for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                        for c in cell:
                            c.fill = mm_fill
            status_col = col_names.index("비교 상태") + 1 if "비교 상태" in col_names else None
            if status_col:
                for row in range(2, ws.max_row + 1):
                    status = str(ws.cell(row=row, column=status_col).value or "")
                    fill = auto_fill if "허용범위" in status or "10% 이내" in status or "비교 가능" in status else warn_fill if "단위" in status or "입력" in status or "10% 초과" in status or "주의" in status else None
                    if fill:
                        ws.cell(row=row, column=status_col).fill = fill
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)
            ws.freeze_panes = "A2"
        paths.append(xlsx_path)
    except Exception:
        pass
    # v0.5.7: for a single current-session export, also create traceable raw/session debug files.
    if session_ids and len(session_ids) == 1:
        try:
            paths.extend(export_session_debug_files(session_ids[0], final_rows=rows))
        except Exception:
            pass
    return paths
